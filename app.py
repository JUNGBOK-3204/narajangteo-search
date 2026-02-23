import os
import time
import requests
import xml.etree.ElementTree as ET
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# ==========================================
# 1. API 키 및 검색 기록 관리 기능
# ==========================================
HISTORY_FILE = "search_history.txt"

def load_api_key():
    """세션에 저장된 키를 확인합니다."""
    if 'user_api_key' in st.session_state:
        return st.session_state['user_api_key']
    return ""

def save_api_key(key):
    """입력받은 키를 세션에 저장합니다."""
    st.session_state['user_api_key'] = key.strip()

def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return [line.strip() for line in f.readlines() if line.strip()]
        except: return []
    return []

def add_history(record):
    history = load_history()
    history.insert(0, record)
    history = history[:5]
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            for item in history:
                f.write(f"{item}\n")
    except: pass

# ==========================================
# 2. API URL 설정
# ==========================================
ORDER_PLAN_URL = "https://apis.data.go.kr/1230000/ao/OrderPlanSttusService/getOrderPlanSttusListServcPPSSrch"
PRIOR_SPEC_URL = "https://apis.data.go.kr/1230000/ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoServcPPSSrch"
BID_NOTICE_URL = "https://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoServcPPSSrch"

def fetch_data_from_api(url, params):
    all_items = []
    page = 1
    
    while True:
        params["pageNo"] = str(page)
        params["numOfRows"] = "500"
            
        try:
            response = requests.get(url, params=params, timeout=30)
            
            if response.status_code != 200: break
            
            root = ET.fromstring(response.text)
            items = root.findall(".//items/item")
            
            if not items: break
            
            for item in items:
                row_data = {child.tag: (child.text or "").strip() for child in list(item)}
                all_items.append(row_data)
            
            total_count_elem = root.find(".//body/totalCount")
            if total_count_elem is not None:
                if len(all_items) >= int(total_count_elem.text): break
            
            page += 1
            
        except Exception as e:
            break
            
    return all_items

def fetch_bid_data_split(service_key, keywords, months=12, progress_callback=None):
    all_results = []
    now = datetime.now()
    for i in range(months):
        start_date = now - timedelta(days=(i + 1) * 30)
        end_date = now - timedelta(days=i * 30)
        bgn_str = start_date.strftime("%Y%m%d%H%M")
        end_str = end_date.strftime("%Y%m%d%H%M")
        for kw in keywords:
            p = {"serviceKey": service_key, "type": "xml", "inqryDiv": "1",
                 "inqryBgnDt": bgn_str, "inqryEndDt": end_str, "bidNtceNm": kw}
            all_results.extend(fetch_data_from_api(BID_NOTICE_URL, p))
        if progress_callback:
            progress_callback(i + 1, months)
        time.sleep(0.1)
    return all_results

# ==========================================
# 3. 데이터 가공 함수
# ==========================================
def get_val(row, possible_keys, default=''):
    row_lower = {k.lower(): v for k, v in row.items()}
    for k in possible_keys:
        if k in row and pd.notna(row[k]) and str(row[k]).strip() != '': return str(row[k]).strip()
        k_lower = k.lower()
        if k_lower in row_lower and pd.notna(row_lower[k_lower]) and str(row_lower[k_lower]).strip() != '': return str(row_lower[k_lower]).strip()
    return default

def get_clean_val(row, possible_keys):
    val = get_val(row, possible_keys)
    return val if val else ""

def apply_exclusion_filter(df, target_col, exclude_keywords):
    if df.empty or not exclude_keywords: return df
    mask = df[target_col].apply(lambda x: not any(exc in str(x) for exc in exclude_keywords))
    return df[mask]

def process_order_for_excel(df, exclude_keywords=[]):
    if df is None or df.empty: return pd.DataFrame()
    new_df = pd.DataFrame()
    new_df['용역구분'] = df.apply(lambda r: '일반용역' if get_val(r, ['bsnsDivCd'])=='03' else ('기술용역' if get_val(r, ['bsnsDivCd'])=='05' else ''), axis=1)
    new_df['업무구분'] = df.apply(lambda r: get_val(r, ['bsnsTyNm', 'bztyNm']), axis=1)
    new_df['발주시기'] = df.apply(lambda r: f"{get_val(r,['orderYear'])}/{get_val(r,['orderMnth']).zfill(2)}" if get_val(r,['orderYear']) else '', axis=1)
    new_df['사업명'] = df.apply(lambda r: get_clean_val(r, ['bizNm', 'prdctClsfNoNm', 'prdctClsfcNoNm', 'cntrctNm']), axis=1)
    new_df['총발주금액(원)'] = df.apply(lambda r: int(float(get_val(r, ['sumOrderAmt', 'totlAmt'], '0').replace(',',''))), axis=1)
    new_df['발주기관명'] = df.apply(lambda r: get_clean_val(r, ['orderInsttNm', 'dmndInsttNm', 'realOrgNm']), axis=1)
    new_df['게시일자'] = df.apply(lambda r: get_val(r, ['nticeDt', 'opengDt']), axis=1)
    new_df = new_df[new_df['사업명'].str.strip() != ''].copy()
    if exclude_keywords:
        new_df = apply_exclusion_filter(new_df, '사업명', exclude_keywords)
    if '발주시기' in new_df.columns:
        new_df.sort_values(by='발주시기', ascending=True, inplace=True)
    new_df.reset_index(drop=True, inplace=True)
    
    # 컬럼 추가 (검토일시 Column A, No. Column B)
    new_df.insert(0, '검토일시', '')
    new_df.insert(1, 'No.', range(1, len(new_df) + 1))
    return new_df

def process_prior_for_excel(df, exclude_keywords=[]):
    if df is None or df.empty: return pd.DataFrame()
    col_map = {
        'bsnsDivNm': '업무구분', 'refNo': '참조번호', 'prdctClsfcNoNm': '사업명(품명)',
        'orderInsttNm': '공고기관', 'rlDminsttNm': '실수요기관', 'asignBdgtAmt': '배정예산액(원)',
        'rcptDt': '공개일시', 'opninRgstClseDt': '의견등록마감일시', 'ofclNm': '담당자명',
        'ofclTelNo': '담당자전화번호', 'swBizObjYn': 'SW사업여부', 'dlvrTmlmtDt': '납품기한',
        'dlvrDaynum': '납품일수', 'bfSpecRgstNo': '사전규격등록번호', 'specDocFileUrl1': '규격서URL',
        'rgstDt': '등록일시', 'chgDt': '변경일시', 'bidNtceNoList': '본공고번호(연계)'
    }
    new_df = pd.DataFrame()
    def parse_money(x):
        try: return int(float(str(x).replace(',', '').strip()))
        except: return 0
    for tag, kr_col in col_map.items():
        found_col = None
        for col in df.columns:
            if col.lower() == tag.lower():
                found_col = col
                break
        if tag == 'asignBdgtAmt':
            new_df[kr_col] = df[found_col].apply(parse_money) if found_col else 0
        else:
            new_df[kr_col] = df[found_col].fillna("") if found_col else ""
    new_df = new_df[new_df['사업명(품명)'].str.strip() != ''].copy()
    if exclude_keywords:
        new_df = apply_exclusion_filter(new_df, '사업명(품명)', exclude_keywords)
    if '공개일시' in new_df.columns:
        new_df.sort_values(by='공개일시', ascending=True, inplace=True)
    new_df.reset_index(drop=True, inplace=True)
    
    # 컬럼 추가 (검토일시 Column A, No. Column B)
    new_df.insert(0, '검토일시', '')
    new_df.insert(1, 'No.', range(1, len(new_df) + 1))
    return new_df

def process_bid_for_excel(df, exclude_keywords=[]):
    if df is None or df.empty: return pd.DataFrame()
    col_map = {
        'bidNtceNo': '공고번호', 'bidNtceOrd': '차수', 'reNtceYn': '재공고여부', 'bidNtceNm': '공고명',
        'ntceKindNm': '공고종류', 'bidMethdNm': '입찰방식', 'cntrctCnclsMthdNm': '계약방법',
        'sucsfbidMthdNm': '낙찰자결정방법', 'ntceInsttNm': '공고기관', 'dminsttNm': '수요기관',
        'ntceInsttOfclNm': '담당자명', 'ntceInsttOfclTelNo': '담당자전화번호', 'bidNtceDt': '게시일시',
        'bidBeginDt': '입찰개시일시', 'bidClseDt': '입찰마감일시', 'opengDt': '개찰일시',
        'bidQlfctRgstDt': '입찰참가자격등록마감일시', 'asignBdgtAmt': '배정예산(원)',
        'presmptPrce': '추정가격(원)', 'bidPrtcptFee': '입찰참가수수료', 'bidNtceUrl': '공고링크(URL)',
        'bfSpecRgstNo': '사전규격등록번호', 'refNo': '참조번호', 'cmmnSpldmdMethdNm': '공동수급여부',
        'prearngPrceDcsnMthdNm': '예가방식', 'opengPlce': '개찰장소', 'brffcBidprcPermsnYn': '지사투찰허용여부'
    }
    new_df = pd.DataFrame()
    def parse_money(x):
        try: return int(float(str(x).replace(',', '').strip()))
        except: return 0
    for tag, kr_col in col_map.items():
        found_col = None
        for col in df.columns:
            if col.lower() == tag.lower():
                found_col = col
                break
        if tag in ['asignBdgtAmt', 'presmptPrce', 'bidPrtcptFee']:
            new_df[kr_col] = df[found_col].apply(parse_money) if found_col else 0
        else:
            new_df[kr_col] = df[found_col].fillna("") if found_col else ""
    def calc_min_bid(row):
        budget = row['배정예산(원)']
        estim = row['추정가격(원)']
        base_price = budget if budget > 0 else estim
        if base_price > 0: return int(base_price * 0.87745)
        return 0
    new_df['예상 투찰하한가(원)'] = new_df.apply(calc_min_bid, axis=1)
    new_df = new_df[new_df['공고명'].str.strip() != ''].copy()
    if exclude_keywords:
        new_df = apply_exclusion_filter(new_df, '공고명', exclude_keywords)
    if '게시일시' in new_df.columns:
        new_df.sort_values(by='게시일시', ascending=True, inplace=True)
    new_df.reset_index(drop=True, inplace=True)
    
    # 컬럼 추가 (검토일시 Column A, No. Column B)
    new_df.insert(0, '검토일시', '')
    new_df.insert(1, 'No.', range(1, len(new_df) + 1))
    return new_df

# ==========================================
# 4. 엑셀 서식화
# ==========================================
def convert_df_to_excel(df_order, df_prior, df_bid):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        align_rules = {
            '발주계획': {'left': ['사업명', '발주기관명'], 'right': ['총발주금액(원)']},
            '사전규격공개': {'left': ['사업명(품명)', '공고기관', '실수요기관'], 'right': ['배정예산액(원)']},
            '입찰공고': {'left': ['공고명', '공고기관', '수요기관'], 'right': ['배정예산(원)', '추정가격(원)', '입찰참가수수료', '예상 투찰하한가(원)']},
        }
        custom_widths = {
            '발주계획': {'검토일시': 15, '사업명': 60}, 
            '사전규격공개': {'검토일시': 15, '사업명(품명)': 60}, 
            '입찰공고': {'검토일시': 15, '공고명': 60, '개찰장소': 32}
        }
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        link_font = Font(color="0000FF", underline="single")
        thin_side = Side(style='thin')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        def apply_styles(df, sheet_name):
            if df is None: return
            if df.empty:
                pd.DataFrame({'결과': ['조건에 맞는 검색 결과가 없습니다.']}).to_excel(writer, index=False, sheet_name=sheet_name)
                return
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            ws.auto_filter.ref = ws.dimensions
            money_cols = ['총발주금액(원)', '배정예산액(원)', '배정예산(원)', '추정가격(원)', '입찰참가수수료', '예상 투찰하한가(원)']
            rules = align_rules.get(sheet_name, {'left': [], 'right': []})
            wide_cols = rules['left']
            sheet_custom_widths = custom_widths.get(sheet_name, {})
            header_map = {}
            for cell in ws[1]: header_map[cell.column] = cell.value
            
            for column in ws.columns:
                column_letter = get_column_letter(column[0].column)
                col_name = header_map.get(column[0].column)
                if col_name in sheet_custom_widths:
                    ws.column_dimensions[column_letter].width = sheet_custom_widths[col_name]
                    continue
                max_length = 0
                for cell in column:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length: max_length = cell_len
                    except: pass
                adjusted_width = (max_length + 7) * 1.4 if col_name in wide_cols else (max_length + 5) * 1.3
                if adjusted_width > 120: adjusted_width = 120
                ws.column_dimensions[column_letter].width = adjusted_width

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border_all
                    col_name = header_map.get(cell.column)
                    if col_name in money_cols and isinstance(cell.value, (int, float)): cell.number_format = '#,##0'
                    if col_name in rules['left']: cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                    elif col_name in rules['right']: cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
                    else: cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if col_name in ['규격서URL', '공고링크(URL)', '공고URL'] and cell.value:
                        val = str(cell.value); cell.hyperlink = val; cell.font = link_font

            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center'); cell.fill = header_fill
                cell.border = border_all

        apply_styles(df_order, '발주계획')
        apply_styles(df_prior, '사전규격공개')
        apply_styles(df_bid, '입찰공고')
    return output.getvalue()

# ==========================================
# 5. UI 및 메인 로직
# ==========================================
st.set_page_config(page_title="나라장터 검색 시스템", layout="wide")

if 'df_order' not in st.session_state: st.session_state.df_order = None
if 'df_prior' not in st.session_state: st.session_state.df_prior = None
if 'df_bid' not in st.session_state: st.session_state.df_bid = None

LOGO_FILENAME = "radsol_logo.png"
col1, col2, col3, col4 = st.columns([1, 6, 1.5, 1.5])
with col1:
    if os.path.exists(LOGO_FILENAME): st.image(LOGO_FILENAME, use_container_width=True)
with col2:
    st.markdown("""
        # 나라장터 검색 시스템 <span style='font-size: 0.5em; color: #cccccc;'>v0.4 &nbsp;&nbsp; by 연구관리팀</span>
        <div style='margin-top: 5px;'>
            <span style='font-size: 15px; color: red; font-weight: bold;'>※ 본 프로그램의 검색 결과는 오류가 발생할 수 있으므로, 중요한 데이터는 꼭 실제 공고를 확인할 것!</span>
        </div>
        """, unsafe_allow_html=True)
with col3:
    st.markdown("<div style='margin-top: 22px;'></div>", unsafe_allow_html=True)
    search_clicked = st.button("🔍 조회 시작", use_container_width=True)
with col4:
    st.markdown("<div style='margin-top: 22px;'></div>", unsafe_allow_html=True)
    download_container = st.empty()

with st.sidebar:
    st.header("⚙️ 검색 설정")
    
    saved_key = load_api_key()
    with st.expander("🔐 내 API Key 설정", expanded=(not saved_key)):
        service_key_input = st.text_input("공공데이터포털 인증키", value=saved_key, type="password", help="본인의 개인 인증키를 입력하세요. 이 키는 세션이 유지되는 동안에만 사용됩니다.")
        if st.button("인증키 적용하기"):
            if service_key_input:
                save_api_key(service_key_input)
                st.success("인증키가 적용되었습니다!")
                st.rerun()
    
    service_key = load_api_key()
    if not service_key: st.error("⛔ 조회를 위해 API 인증키 입력이 필요합니다.")
    
    st.divider()
    st.subheader("📋 조회 대상 선택")
    
    check_order = st.checkbox("발주계획 (일반/기술용역)", value=True, help="나라장터 발주계획 중 '일반용역'과 '기술용역' 분야를 조회합니다.")
    check_prior = st.checkbox("사전규격공개 (일반/기술용역)", value=True, help="사전규격공개 중 '일반용역'과 '기술용역' 분야를 조회합니다.")
    check_bid = st.checkbox("입찰공고 (일반/기술용역)", value=True, help="입찰공고 중 '일반용역'과 '기술용역' 분야를 조회합니다.")
    
    st.divider()

    keywords_input = st.text_input(
        "🔑 키워드 (쉼표로 구분)", 
        value="방사능", 
        help="검색하고 싶은 핵심 단어를 입력하세요. 쉼표(,)로 구분하여 여러 개를 동시에 검색할 수 있습니다. (예: 방사능, 원자력, 폐기물)"
    )
    keywords = [k.strip() for k in keywords_input.split(",") if k.strip()]
    
    exclude_input = st.text_input(
        "🚫 제외 키워드 (쉼표로 구분)", 
        value="유지보수, X-ray", 
        help="검색 결과 중에서 이 단어가 포함된 공고는 제외합니다. (예: 유지보수, 단순구매)"
    )
    exclude_keywords = [k.strip() for k in exclude_input.split(",") if k.strip()]
    
    st.divider()
    
    year = int(st.number_input(
        "조회 연도 설정", 
        min_value=2000, 
        value=2026, 
        help="검색 기준 연도를 설정합니다. [발주계획]은 '발주예정시기'가 해당 연도인 건을, [입찰/사전규격]은 '공고게시일'이 해당 연도(1.1~12.31)인 건을 조회합니다."
    ))
    
    bid_months = st.slider(
        "입찰공고 수집 기간 (최근 N개월)", 
        1, 12, 3, 
        disabled=not check_bid,
        help="입찰공고는 데이터량이 많아 최근 N개월치만 수집합니다. 기간이 길수록 조회 시간이 늘어날 수 있습니다."
    )
    
    st.divider()
    history_container = st.empty()
    def update_history_ui():
        with history_container.container():
            st.subheader("🕒 최근 조회 (공용)")
            for h in load_history(): st.caption(f"• {h}")
    update_history_ui()

if search_clicked:
    if not service_key: 
        st.warning("먼저 사이드바에서 API 인증키를 입력하고 [적용하기]를 눌러주세요.")
    else:
        st.session_state.df_order = None
        st.session_state.df_prior = None
        st.session_state.df_bid = None
        
        prog_bar = st.progress(0, text="데이터 수집 시작...")
        inqry_bgn, inqry_end = f"{year}01010000", f"{year}12312359"
        total_kw = len(keywords)
        
        if check_order or check_prior:
            all_o, all_p = [], []
            for idx, kw in enumerate(keywords):
                current_progress = int(40 * (idx + 1) / (total_kw if total_kw > 0 else 1))
                prog_bar.progress(current_progress, text=f"🔍 나라장터 조회 중... ({kw})")
                if check_order:
                    for cd in ["03", "05"]:
                        all_o.extend(fetch_data_from_api(ORDER_PLAN_URL, {"serviceKey": service_key, "type": "xml", "inqryBgnDt": inqry_bgn, "inqryEndDt": inqry_end, "orderBgnYm": f"{year-1}12", "orderEndYm": f"{year}12", "bsnsDivCd": cd, "bizNm": kw}))
                if check_prior:
                    all_p.extend(fetch_data_from_api(PRIOR_SPEC_URL, {"serviceKey": service_key, "type": "xml", "inqryDiv": "1", "inqryBgnDt": inqry_bgn, "inqryEndDt": inqry_end, "prdctClsfcNoNm": kw}))
            if check_order: st.session_state.df_order = process_order_for_excel(pd.DataFrame(all_o).drop_duplicates(), exclude_keywords)
            if check_prior: st.session_state.df_prior = process_prior_for_excel(pd.DataFrame(all_p).drop_duplicates(), exclude_keywords)

        if check_bid:
            def bid_progress_update(current, total):
                percent = 40 + int(60 * (current / total))
                prog_bar.progress(min(percent, 100), text=f"📥 입찰공고 수집 중... ({current}/{total}개월)")
            all_b = fetch_bid_data_split(service_key, keywords, months=bid_months, progress_callback=bid_progress_update)
            st.session_state.df_bid = process_bid_for_excel(pd.DataFrame(all_b).drop_duplicates(), exclude_keywords)
        
        prog_bar.progress(100, text="✅ 완료!")
        time.sleep(0.5); prog_bar.empty()
        
        cnt_o = len(st.session_state.df_order) if st.session_state.df_order is not None else 0
        cnt_p = len(st.session_state.df_prior) if st.session_state.df_prior is not None else 0
        cnt_b = len(st.session_state.df_bid) if st.session_state.df_bid is not None else 0
        st.success(f"✅ 조회가 완료되었습니다! [ 발주: {cnt_o}건 / 사전: {cnt_p}건 / 입찰: {cnt_b}건 ]")

        add_history(f"{datetime.now().strftime('%m/%d %H:%M:%S')} ({keywords_input})")
        update_history_ui()

if any(x is not None for x in [st.session_state.df_order, st.session_state.df_prior, st.session_state.df_bid]):
    xl_data = convert_df_to_excel(st.session_state.df_order, st.session_state.df_prior, st.session_state.df_bid)
    download_container.download_button(label="📥 엑셀 다운로드", data=xl_data, file_name=f"통합조회_{datetime.now().strftime('%Y%m%d')}.xlsx", use_container_width=True)
    
    tabs_labels = []
    if st.session_state.df_order is not None: tabs_labels.append("📊 발주계획")
    if st.session_state.df_prior is not None: tabs_labels.append("📝 사전규격공개")
    if st.session_state.df_bid is not None: tabs_labels.append("🔔 입찰공고")
    
    if tabs_labels:
        tabs = st.tabs(tabs_labels)
        curr = 0
        if st.session_state.df_order is not None:
            with tabs[curr]:
                df_v = st.session_state.df_order.copy()
                if not df_v.empty: df_v['총발주금액(원)'] = df_v['총발주금액(원)'].apply(lambda x: f"{x:,}")
                st.dataframe(df_v, use_container_width=True, hide_index=True, height=600)
            curr += 1
        if st.session_state.df_prior is not None:
            with tabs[curr]:
                df_p = st.session_state.df_prior.copy()
                if not df_p.empty: df_p['배정예산액(원)'] = df_p['배정예산액(원)'].apply(lambda x: f"{x:,}" if x > 0 else "")
                st.dataframe(df_p, use_container_width=True, hide_index=True, height=600, column_config={"규격서URL": st.column_config.LinkColumn("규격서", display_text="🔗 다운로드")})
            curr += 1
        if st.session_state.df_bid is not None:
            with tabs[curr]:
                df_b = st.session_state.df_bid.copy()
                for col in ['배정예산(원)', '추정가격(원)', '입찰참가수수료', '예상 투찰하한가(원)']:
                    if not df_b.empty and col in df_b.columns: df_b[col] = df_b[col].apply(lambda x: f"{x:,}" if x > 0 else "")
                st.dataframe(df_b, use_container_width=True, hide_index=True, height=600, column_config={"공고링크(URL)": st.column_config.LinkColumn("원문 링크", display_text="🔗 공고이동")})
            curr += 1
